#!/usr/bin/env python
"""Extract Israel-relevant foundations from the xlsx into a JSON payload for the verifier HTML."""
import sys, io, json, html, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

SRC = r'c:\Users\meyta\Downloads\קרנות_פילנתרופיות.xlsx'
OUT_JSON = r'c:\Users\meyta\Downloads\impactos\_handoff\foundations-data.json'

wb = openpyxl.load_workbook(SRC, data_only=True)

# ---- Sheet 1: main ----
ws = wb['קרנות פילנתרופיות']
# columns: שם, טלפון, אימייל, אתר, איש קשר, תחום, תיאור, סוג, סטטוס, דרך פנייה, תבחינים, רלוונטי
COLS = ['name','phone','email','website','contact','area','description','source','status','how_to_apply','criteria','relevance']

# Track section context based on divider rows
def is_divider(row):
    if not row or not row[0]:
        return False
    s = str(row[0])
    return '▼' in s

def section_name(row):
    s = str(row[0]).replace('▼','').strip()
    # strip count "(N)"
    import re
    s = re.sub(r'\(\d+\)', '', s).strip()
    return s

# Sections to SKIP (not relevant to Israel)
SKIP_PATTERNS = [
    'שלא תורמות לישראל',
    'לא רלוונטיים',
    'לא לפנות',
]

current_section = None
keep_current = True
main_rows = []
section_counts = {}

for row in ws.iter_rows(min_row=2, values_only=True):
    if is_divider(row):
        current_section = section_name(row)
        keep_current = not any(p in current_section for p in SKIP_PATTERNS)
        continue
    # also filter by per-row relevance flag (col 11)
    relevance = row[11] if len(row) > 11 else None
    if not keep_current:
        continue
    # skip totally empty rows
    if not any(row):
        continue
    # skip rows where relevance is explicitly "❌ לא רלוונטי"
    if relevance and 'לא רלוונטי' in str(relevance):
        continue
    item = {COLS[i]: row[i] for i in range(min(len(COLS), len(row)))}
    item['_section'] = current_section
    item['_sheet'] = 'main'
    if item.get('name'):
        main_rows.append(item)
        section_counts[current_section] = section_counts.get(current_section, 0) + 1

print('=== Main sheet sections kept ===')
for k, v in section_counts.items():
    print(f'  {k}: {v}')
print(f'Total main: {len(main_rows)}')

# ---- Sheet 2: Atlas - Foundations ----
ws2 = wb['אטלס - קרנות']
ATLAS_COLS = ['idx','name','year','founders','focus','target','tags','financial','israeli','donates_to_israel','duplicate','phone','email','website','contact','research_status']
atlas_rows = []
for row in ws2.iter_rows(min_row=2, values_only=True):
    if not row or not row[1]:
        continue
    donates = row[9] if len(row) > 9 else None
    # skip those that don't donate
    if donates and 'לא' in str(donates) and 'כן' not in str(donates):
        continue
    item = {ATLAS_COLS[i]: row[i] for i in range(min(len(ATLAS_COLS), len(row)))}
    # skip duplicates already in main sheet
    dup = str(item.get('duplicate') or '')
    if 'כן' in dup:
        # mark but still include — let user decide
        item['_is_duplicate'] = True
    item['_section'] = 'אטלס - קרנות'
    item['_sheet'] = 'atlas'
    atlas_rows.append(item)
print(f'Total atlas-foundations: {len(atlas_rows)}')

# ---- Sheet 3: Atlas - Endowments (Israeli) ----
ws3 = wb['אטלס - הקדשים']
END_COLS = ['idx','name','purpose','assets','year','categories','trustee','contact_info','how_to_apply','other_sources','research_status']
endow_rows = []
for row in ws3.iter_rows(min_row=2, values_only=True):
    if not row or not row[1]:
        continue
    item = {END_COLS[i]: row[i] for i in range(min(len(END_COLS), len(row)))}
    item['_section'] = 'אטלס - הקדשים ישראליים'
    item['_sheet'] = 'endowments'
    endow_rows.append(item)
print(f'Total endowments: {len(endow_rows)}')

all_items = []
for i, r in enumerate(main_rows + atlas_rows + endow_rows):
    r['_id'] = f'r{i}'
    # normalize None → '' for cleaner JSON
    for k, v in list(r.items()):
        if v is None:
            r[k] = ''
        else:
            r[k] = str(v).strip()
    all_items.append(r)

print(f'TOTAL items for verification: {len(all_items)}')

# write JSON
with open(OUT_JSON, 'w', encoding='utf-8') as f:
    json.dump(all_items, f, ensure_ascii=False, indent=2)
print(f'Wrote {OUT_JSON}')
